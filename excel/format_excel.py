from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, numbers

wb = load_workbook('sales.xlsx')
ws = wb.active
# for cell in ws[2]:
#     cell.font = Font(bold=True)
# wb.save('fonted_excel.xlsx')
# green_fill = PatternFill(start_color='008000',end_color='008000', fill_type='solid')
# for row in ws.iter_rows(min_row=2, max_col=7, max_row=ws.max_row):
#     # print(row)
#     # break
#     if row[6].value > 500:
#         for cell in row:
#             cell.fill = green_fill

# print(ws['A']) # month/day/year


# for cell in ws['A']:
#     if cell.row != 1:
#         cell.number_format = numbers.FORMAT_DATE_XLSX22


ws.freeze_panes = 'E7'

wb.save('fonted_excel.xlsx')
